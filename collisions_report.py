import openpyxl
from openpyxl.styles import Border, Side, Alignment, Font
import tkinter as tk
from tkinter import filedialog, simpledialog, messagebox
from openpyxl.utils import column_index_from_string


HEADER_ROW = 7
NUMBER_ROW = 8


def format_excel_report():
    # Выбор файла
    filename = filedialog.askopenfilename(
        title="Выберите Excel файл",
        filetypes=[("Excel files", "*.xlsx *.xlsm")]
    )
    if not filename:
        return

    try:
        wb = openpyxl.load_workbook(filename)
    except Exception as e:
        messagebox.showerror("Ошибка", f"Не удалось открыть файл:\n{e}")
        return

    # Выбор листа
    if len(wb.sheetnames) == 1:
        sheet_name = wb.sheetnames[0]
    else:
        sheet_name = simpledialog.askstring(
            "Выбор листа",
            f"Доступные листы:\n{', '.join(wb.sheetnames)}\n\nВведите название листа:"
        )
        if not sheet_name or sheet_name not in wb.sheetnames:
            messagebox.showerror("Ошибка", "Такого листа нет!")
            return

    ws = wb[sheet_name]

    # Удаляем остальные листы
    for sheet in wb.sheetnames:
        if sheet != sheet_name:
            del wb[sheet]

    # Вставляем первый столбец "Конфликт"
    ws.insert_cols(1)
    ws.cell(row=HEADER_ROW, column=1, value="Конфликт")

    # Определяем номера колонок
    headers = {ws.cell(row=HEADER_ROW, column=col).value: col for col in range(1, ws.max_column + 1)}
    id1_col = headers.get("ID 1го")
    id2_col = headers.get("ID 2го")

    if not id1_col or not id2_col:
        messagebox.showerror("Ошибка", "Не найдены столбцы 'ID 1го' и 'ID 2го'!")
        return

    # Нумерация конфликтов
    conflict_counter = 0
    for row in range(HEADER_ROW + 1, ws.max_row + 1):
        id1 = ws.cell(row=row, column=id1_col).value
        id2 = ws.cell(row=row, column=id2_col).value

        if not id1 and not id2:
            conflict_counter = 0
            ws.cell(row=row, column=1, value=None)
        elif id1 and id2 and row > HEADER_ROW + 1:
            conflict_counter += 1
            ws.cell(row=row, column=1, value=f"Конфликт {conflict_counter}")
        else:
            ws.cell(row=row, column=1, value=None)

    # Добавляем последний столбец "Комментарий ПТИ"
    last_col = ws.max_column + 1
    ws.cell(row=HEADER_ROW, column=last_col, value="Комментарий ПТИ")
    ws.cell(row=HEADER_ROW + 1, column=last_col, value=12)

    # Переименование столбца "Комментарий" → "Комментарий BIM отдела"
    for col in range(1, ws.max_column + 1):
        if ws.cell(row=HEADER_ROW, column=col).value == "Комментарий":
            ws.cell(row=HEADER_ROW, column=col, value="Комментарий BIM отдела")

    # Скрываем столбцы E, H, I, J, K
    for col_letter in ["E", "H", "I", "J", "K"]:
        ws.column_dimensions[col_letter].hidden = True

    # Границы
    thin = Side(border_style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for row in ws.iter_rows(min_row=HEADER_ROW, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.border = border

    # Заголовки и номера
    header_font = Font(bold=True)
    align_center = Alignment(horizontal="center", vertical="center")

    for col in range(1, ws.max_column + 1):
        ws.cell(row=HEADER_ROW, column=col).font = header_font
        ws.cell(row=HEADER_ROW, column=col).alignment = align_center
        ws.cell(row=NUMBER_ROW, column=col).alignment = align_center

    # === Коррекция ширины столбцов ===
    # Меняем C↔D
    col_C = ws.column_dimensions['C']
    col_D = ws.column_dimensions['D']
    width_C = col_C.width if col_C.width else 10
    width_D = col_D.width if col_D.width else 10
    col_C.width = width_D * 0.8  # уменьшаем C на 20%
    col_D.width = width_C
    # Меняем F↔G
    col_F = ws.column_dimensions['F']
    col_G = ws.column_dimensions['G']
    width_F = col_F.width if col_F.width else 10
    width_G = col_G.width if col_G.width else 10
    col_F.width = width_G * 0.8  # уменьшаем F на 20%
    col_G.width = width_F
    # Увеличиваем L и M в 3 раза
    for col_letter in ['L', 'M']:
        col = ws.column_dimensions[col_letter]
        col.width = (col.width if col.width else 10) * 3

    # Вставляем невидимый символ в L, строки 1–6
    col_L = column_index_from_string('L')
    for row in range(1, 5):
        cell = ws.cell(row=row, column=col_L)
        if isinstance(cell, openpyxl.cell.cell.MergedCell):
            for merged_range in ws.merged_cells.ranges:
                if cell.coordinate in merged_range:
                    cell = ws.cell(row=merged_range.min_row, column=merged_range.min_col)
                    break
        cell.value = "•"
        cell.font = Font(color="FFFFFF")

    # === Сохраняем файл с именем из B3 ===
    name_cell = ws.cell(row=3, column=3).value
    if name_cell:
        safe_name = "".join(c for c in str(name_cell) if c not in r'\/:*?"<>|')
        new_filename = filename.rsplit("/", 1)[0] + f"/{safe_name}.xlsx"
    else:
        new_filename = filename.replace(".xlsx", "_formatted.xlsx")

    wb.save(new_filename)
    messagebox.showinfo("Готово", f"Файл сохранен:\n{new_filename}")


# GUI
root = tk.Tk()
root.title("Форматирование Excel отчета")

frame = tk.Frame(root, padx=20, pady=20)
frame.pack()

label = tk.Label(frame, text="Выберите Excel файл для форматирования:")
label.pack(pady=5)

button = tk.Button(frame, text="Открыть файл", command=format_excel_report)
button.pack(pady=10)

root.mainloop()
